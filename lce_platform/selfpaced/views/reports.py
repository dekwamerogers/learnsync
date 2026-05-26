import io
from collections import defaultdict
from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Min, OuterRef, Q, Subquery
from django.db.models.functions import Coalesce, Greatest, TruncWeek
from django.http import HttpResponse
from django.shortcuts import render

from selfpaced.models import Course, CourseEnrolment, Enrolment, HealthStatus, Intervention, Learner, PaymentStatus

# ── Lookup tables ──────────────────────────────────────────────────────────────

_FLAG_LABELS = [
    ('inactive',                'Inactive'),
    ('never_activated',         'Never Activated'),
    ('stuck_on_assignment',     'Stuck on Assignment'),
    ('low_pass_rate',           'Low Pass Rate'),
    ('stalled_between_courses', 'Stalled Between Courses'),
    ('stalled_progression',     'No Onward Progress'),
    ('payment_issue',           'Payment Issue'),
]

_FLAG_DESCRIPTIONS = {
    'inactive':                "No assignment activity for longer than the inactivity threshold (typically 7 days). The learner was active before but has gone quiet.",
    'never_activated':         "The learner is enrolled but has not yet passed the first module of their programme. They have not made it past the starting point.",
    'stuck_on_assignment':     "The learner has been on the same assignment for longer than the stuck threshold without submitting. May need a nudge.",
    'low_pass_rate':           "The learner's pass rate on submitted assignments has dropped below the acceptable threshold. May need academic support.",
    'stalled_between_courses': "The learner finished a course but hasn't started the next one within the expected window. May need encouragement to continue.",
    'stalled_progression':     "The learner is active but not making measurable progress toward graduation. Activity without advancement.",
    'payment_issue':           "There is an open payment concern for this learner — they may be overdue or in a grace period.",
}

_HEALTH_DISPLAY = {
    'active':          'Active',
    'at_risk':         'At Risk',
    'dormant':         'Dormant',
    'graduated':       'Graduated',
    'not_yet_started': 'Not Started',
}

# Metric definitions shown as tooltips on the page and in the Excel definitions sheet
METRIC_DEFINITIONS = {
    'Paid Enrolled':     "Learners with an active or pending payment who are enrolled in at least one programme. Learners with no payment on record (status: Unpaid) are excluded from every metric on this page.",
    'Activated':         "Learners who have passed the first module of their programme — the point at which a learner moves from 'just enrolled' to 'actively participating.'",
    'Activation Rate':   "Activated ÷ Enrolled × 100. The share of paid enrolled learners who have passed the first module. A low rate may indicate onboarding friction or engagement barriers.",
    'Retained':          "Learners who passed the first module and remain engaged — health status is Active, At Risk, or Graduated. They have continued beyond the starting point and not disengaged.",
    'Retention Rate':    "Retained ÷ Activated × 100. Of learners who passed the first module, what percentage are still going? A high rate means starters tend to stick with it.",
    'Graduated':         "Learners who completed all required courses and met the graduation criteria. This is the end goal.",
    'Graduation Rate':   "Graduated ÷ Enrolled × 100. The share of enrolled (paid) learners who have reached the finish line. Compare across programmes with caution — older programmes naturally have higher rates.",
    'Badges Earned':     "Course-completion credentials awarded across all programmes. One badge is earned each time a learner passes a course in a credential-awarding programme. Total badges can exceed total learners because each learner can earn multiple.",
    'Active':            "Learners who are on track: submitting work, passing assignments, and within all inactivity thresholds.",
    'At Risk':           "Learners with at least one warning flag raised. See the 'At-Risk Flags' panel for the specific reasons.",
    'Dormant':           "Learners who enrolled or activated but have shown no meaningful activity for an extended period. These learners need re-engagement outreach.",
    'Not Started':       "Learners who are enrolled (and paid) but have not yet shown any activity. They have neither activated nor gone dormant.",
    'Inactive':          _FLAG_DESCRIPTIONS['inactive'],
    'Never Activated':   _FLAG_DESCRIPTIONS['never_activated'],
    'Stuck on Assignment': _FLAG_DESCRIPTIONS['stuck_on_assignment'],
    'Low Pass Rate':     _FLAG_DESCRIPTIONS['low_pass_rate'],
    'Stalled Between Courses': _FLAG_DESCRIPTIONS['stalled_between_courses'],
    'No Onward Progress': _FLAG_DESCRIPTIONS['stalled_progression'],
    'Payment Issue':     _FLAG_DESCRIPTIONS['payment_issue'],
}


def _build_report_data():
    today = date.today()
    thirty_ago = today - timedelta(days=30)

    # Base queryset: active, started, non-prerequisite programmes, PAID learners only.
    # Exclude upcoming programmes (start_date in the future) — they have no learner
    # activity yet and inflate "enrolled" counts misleadingly.
    base_qs = (
        Enrolment.objects
        .filter(programme__is_prerequisite=False, programme__is_active=True, has_activity_data=True)
        .filter(Q(programme__start_date__isnull=True) | Q(programme__start_date__lte=today))
        .exclude(learner__payment_status=PaymentStatus.UNKNOWN)
    )

    # Activated = passed the first module (lowest sequence_number course) of their programme.
    # Build the set of qualifying enrolment PKs from CourseEnrolment data.
    _activated_ids = frozenset(
        CourseEnrolment.objects
        .filter(
            enrolment__in=base_qs,
            is_passed=True,
            course__sequence_number=Subquery(
                Course.objects.filter(
                    programme_id=OuterRef('enrolment__programme_id'),
                ).exclude(code='WALX').order_by('sequence_number').values('sequence_number')[:1]
            ),
        )
        .values_list('enrolment_id', flat=True)
        .distinct()
    )

    prog_rows = list(
        base_qs
        .values('programme__id', 'programme__code', 'programme__name')
        .annotate(
            total=Count('pk'),
            activated=Count('pk', filter=Q(pk__in=_activated_ids)),
            # Retained = passed first module AND still meaningfully engaged
            retained=Count('pk', filter=Q(
                pk__in=_activated_ids,
                health_status__in=[HealthStatus.ACTIVE, HealthStatus.AT_RISK, HealthStatus.GRADUATED],
            )),
            graduated=Count('pk', filter=Q(is_graduated=True)),
            active=Count('pk',      filter=Q(health_status=HealthStatus.ACTIVE)),
            at_risk=Count('pk',     filter=Q(health_status=HealthStatus.AT_RISK)),
            dormant=Count('pk',     filter=Q(health_status=HealthStatus.DORMANT)),
            not_started=Count('pk', filter=Q(health_status=HealthStatus.NOT_YET_STARTED)),
        )
        .order_by('programme__code')
    )

    for p in prog_rows:
        total = p['total']
        act   = p['activated']
        # Activation rate: passed M1 ÷ enrolled
        p['activation_rate'] = round(act / total * 100, 1) if total else 0.0
        # Retention rate: retained ÷ activated
        p['retention_rate']  = round(p['retained'] / act * 100, 1) if act else 0.0
        # Graduation rate: graduated ÷ activated (of those who started, how many finished?)
        p['grad_rate']       = round(p['graduated'] / act * 100, 1) if act else 0.0

    # Badges: passed CourseEnrolments from credential-awarding programmes (paid learners only)
    badges_by_prog = {
        row['enrolment__programme_id']: row['n']
        for row in (
            CourseEnrolment.objects
            .filter(
                is_passed=True,
                course__programme__awards_credentials=True,
                enrolment__programme__is_prerequisite=False,
            )
            .exclude(enrolment__learner__payment_status=PaymentStatus.UNKNOWN)
            .values('enrolment__programme_id')
            .annotate(n=Count('pk'))
        )
    }
    for p in prog_rows:
        p['badges'] = badges_by_prog.get(p['programme__id'], 0)

    # Payment distribution across ALL learners (including unpaid — useful context for managers)
    payment_rows = list(
        Learner.objects
        .values('payment_status')
        .annotate(count=Count('email'))
        .order_by('payment_status')
    )
    total_learners = sum(p['count'] for p in payment_rows)
    label_map = dict(PaymentStatus.choices)
    for p in payment_rows:
        p['pct']   = round(p['count'] / total_learners * 100, 1) if total_learners else 0.0
        p['label'] = label_map.get(p['payment_status'], p['payment_status'])

    # Count of unpaid (excluded) learners so the page can show context
    unpaid_count = next(
        (p['count'] for p in payment_rows if p['payment_status'] == PaymentStatus.UNKNOWN), 0
    )

    # Aggregate totals
    total_enrolled  = sum(p['total']     for p in prog_rows)
    total_activated = sum(p['activated'] for p in prog_rows)
    total_retained  = sum(p['retained']  for p in prog_rows)
    total_graduated = sum(p['graduated'] for p in prog_rows)
    total_badges    = sum(p['badges']    for p in prog_rows)

    # Health totals — paid learners only
    health_totals = {
        row['health_status']: row['n']
        for row in base_qs.values('health_status').annotate(n=Count('pk'))
    }

    # At-risk flag breakdown — paid learners only
    flag_learners: dict = defaultdict(set)
    for row in (
        base_qs
        .filter(health_status__in=[HealthStatus.AT_RISK, HealthStatus.DORMANT])
        .values('learner_id', 'active_flags')
    ):
        for flag in (row['active_flags'] or []):
            flag_learners[flag].add(row['learner_id'])
    flag_rows = [
        {
            'code':  code,
            'flag':  label,
            'count': len(flag_learners.get(code, set())),
            'desc':  _FLAG_DESCRIPTIONS.get(code, ''),
        }
        for code, label in _FLAG_LABELS
    ]

    # Recent interventions (last 30 days, all learners regardless of payment)
    recent_interventions = list(
        Intervention.objects
        .filter(intervention_date__gte=thirty_ago)
        .select_related('learner', 'enrolment__programme', 'logged_by')
        .order_by('-intervention_date')[:200]
    )

    # ── Weekly breakdown (last 13 weeks) ──────────────────────────────────────
    _current_monday = today - timedelta(days=today.weekday())
    _week_starts    = [_current_monday - timedelta(weeks=i) for i in range(12, -1, -1)]
    _lookback       = _week_starts[0]

    def _as_date(v):
        return v.date() if hasattr(v, 'hour') else v

    # Effective enrolment date: the later of enrolment_date and programme start_date.
    # Learners added before their programme launched are attributed to the launch week.
    _eff_date = Greatest(
        Coalesce('enrolment_date', 'programme__start_date'),
        Coalesce('programme__start_date', 'enrolment_date'),
    )

    def _enrolled_qs(qs):
        return (
            qs
            .annotate(_eff=_eff_date)
            .filter(_eff__gte=_lookback, _eff__isnull=False)
            .annotate(week=TruncWeek('_eff'))
        )

    # ── Overall weekly totals ─────────────────────────────────────────────────
    _w_enrolled = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs).values('week').annotate(n=Count('pk'))
    )}
    # Cohort-based activated: of learners whose effective start falls in this week,
    # how many have EVER passed their first module?  Uses the same _activated_ids
    # frozenset built above so the definition is identical to the summary cards.
    # This guarantees Activated ≤ Enrolled for every row.
    _w_activated = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs.filter(pk__in=_activated_ids))
        .values('week').annotate(n=Count('pk'))
    )}
    # Cohort-based: of learners enrolled in this week, how many have graduated?
    # Matches the Active / At Risk approach — bucketed by effective enrolment week,
    # not by graduation_date — so Graduated will never exceed Enrolled.
    _w_graduated = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.GRADUATED))
        .values('week').annotate(n=Count('pk'))
    )}
    _w_interventions = {_as_date(r['week']): r['n'] for r in (
        Intervention.objects
        .filter(intervention_date__gte=_lookback)
        .annotate(week=TruncWeek('intervention_date'))
        .values('week').annotate(n=Count('id'))
    )}
    _w_active = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.ACTIVE))
        .values('week').annotate(n=Count('pk'))
    )}
    _w_at_risk = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.AT_RISK))
        .values('week').annotate(n=Count('pk'))
    )}
    # Cohort-based badges: total passed CourseEnrolments for learners enrolled this week
    _w_badges = {_as_date(r['week']): r['n'] for r in (
        _enrolled_qs(base_qs)
        .values('week')
        .annotate(n=Count(
            'course_enrolments',
            filter=Q(course_enrolments__is_passed=True),
        ))
    )}

    _WEEK_COLS = ('enrolled', 'activated', 'active', 'at_risk', 'graduated', 'badges', 'interventions')
    weekly_rows = [
        row for row in (
            {
                'week_start':    ws,
                'week_end':      ws + timedelta(days=6),
                'label':         f'{ws.strftime("%d %b")} – {(ws + timedelta(days=6)).strftime("%d %b")}',
                'enrolled':      _w_enrolled.get(ws, 0),
                'activated':     _w_activated.get(ws, 0),
                'active':        _w_active.get(ws, 0),
                'at_risk':       _w_at_risk.get(ws, 0),
                'graduated':     _w_graduated.get(ws, 0),
                'badges':        _w_badges.get(ws, 0),
                'interventions': _w_interventions.get(ws, 0),
            }
            for ws in _week_starts
        )
        if any(row[c] for c in _WEEK_COLS)
    ]

    # ── Per-programme per-week breakdown ──────────────────────────────────────
    _prog_name_map = {p['programme__code']: p['programme__name'] for p in prog_rows}

    _pw_enrolled  = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs)
        .values('programme__code', 'week').annotate(n=Count('pk'))
    ):
        _pw_enrolled[r['programme__code']][_as_date(r['week'])] += r['n']

    # Cohort-based: same week bucket as enrolled, same _activated_ids definition
    _pw_activated = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs.filter(pk__in=_activated_ids))
        .values('programme__code', 'week').annotate(n=Count('pk'))
    ):
        _pw_activated[r['programme__code']][_as_date(r['week'])] += r['n']

    _pw_active = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.ACTIVE))
        .values('programme__code', 'week').annotate(n=Count('pk'))
    ):
        _pw_active[r['programme__code']][_as_date(r['week'])] += r['n']

    _pw_at_risk = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.AT_RISK))
        .values('programme__code', 'week').annotate(n=Count('pk'))
    ):
        _pw_at_risk[r['programme__code']][_as_date(r['week'])] += r['n']

    # Cohort-based: same approach as active/at_risk — bucketed by enrolment week
    _pw_graduated = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs.filter(health_status=HealthStatus.GRADUATED))
        .values('programme__code', 'week').annotate(n=Count('pk'))
    ):
        _pw_graduated[r['programme__code']][_as_date(r['week'])] += r['n']

    # Cohort-based badges: total passed CourseEnrolments for learners enrolled this week
    _pw_badges = defaultdict(lambda: defaultdict(int))
    for r in (
        _enrolled_qs(base_qs)
        .values('programme__code', 'week')
        .annotate(n=Count(
            'course_enrolments',
            filter=Q(course_enrolments__is_passed=True),
        ))
    ):
        _pw_badges[r['programme__code']][_as_date(r['week'])] += r['n']

    _pw_interventions = defaultdict(lambda: defaultdict(int))
    for r in (
        Intervention.objects
        .filter(intervention_date__gte=_lookback, enrolment__isnull=False)
        .annotate(week=TruncWeek('intervention_date'))
        .values('enrolment__programme__code', 'week').annotate(n=Count('id'))
    ):
        _pw_interventions[r['enrolment__programme__code']][_as_date(r['week'])] += r['n']

    _all_pw_codes = sorted(
        set(_pw_enrolled) | set(_pw_activated) | set(_pw_active)
        | set(_pw_at_risk) | set(_pw_graduated) | set(_pw_badges) | set(_pw_interventions)
    )
    weekly_by_prog = [
        {
            'code': code,
            'name': _prog_name_map.get(code, code),
            'weeks': [
                w for w in (
                    {
                        'label':         f'{ws.strftime("%d %b")} – {(ws + timedelta(days=6)).strftime("%d %b")}',
                        'enrolled':      _pw_enrolled[code].get(ws, 0),
                        'activated':     _pw_activated[code].get(ws, 0),
                        'active':        _pw_active[code].get(ws, 0),
                        'at_risk':       _pw_at_risk[code].get(ws, 0),
                        'graduated':     _pw_graduated[code].get(ws, 0),
                        'badges':        _pw_badges[code].get(ws, 0),
                        'interventions': _pw_interventions[code].get(ws, 0),
                    }
                    for ws in _week_starts
                )
                if any(w[c] for c in _WEEK_COLS)
            ],
        }
        for code in _all_pw_codes
    ]

    # Learner detail for Excel — paid learners only
    learner_rows = list(
        base_qs
        .order_by('programme__code', 'learner__last_name', 'learner__first_name')
        .values(
            'learner__first_name',
            'learner__last_name',
            'learner__email',
            'learner__country',
            'learner__payment_status',
            'programme__code',
            'programme__name',
            'health_status',
            'current_course__full_name',
            'activation_date',
            'enrolment_date',
            'is_graduated',
            'graduation_date',
        )
    )

    return {
        'today':           today,
        'thirty_ago':      thirty_ago,
        # Summary totals
        'active_programmes':  len(prog_rows),
        'total_enrolled':     total_enrolled,
        'total_activated':    total_activated,
        'total_retained':     total_retained,
        'total_graduated':    total_graduated,
        'total_badges':       total_badges,
        'total_learners':     total_learners,
        'unpaid_count':       unpaid_count,
        # Rates
        'activation_rate':   round(total_activated  / total_enrolled  * 100, 1) if total_enrolled  else 0.0,
        'retention_rate':    round(total_retained   / total_activated * 100, 1) if total_activated else 0.0,
        # Graduation rate: of those who passed Module 1, how many finished?
        'grad_rate':         round(total_graduated  / total_activated * 100, 1) if total_activated else 0.0,
        # Detail rows
        'prog_rows':          prog_rows,
        'payment_rows':       payment_rows,
        'health_totals':      health_totals,
        'flag_rows':          flag_rows,
        'recent_interventions': recent_interventions,
        'weekly_rows':        weekly_rows,
        'weekly_by_prog':     weekly_by_prog,
        'learner_rows':       learner_rows,
        'metric_definitions': METRIC_DEFINITIONS,
    }


@login_required
def manager_report(request):
    fmt = request.GET.get('format', '')
    if fmt == 'excel':
        return _excel_response(_build_report_data())
    if fmt == 'pdf':
        return _pdf_response(_build_report_data())
    data = _build_report_data()
    return render(request, 'selfpaced/manager_report.html', data)


def _excel_response(data):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    BLUE, WHITE, LGRAY, YLLOW = '0452F0', 'FFFFFF', 'F3F4F6', 'FFF9C4'

    def _hdr(cell, bg=BLUE):
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_sheet_rows(ws, rows, start=2):
        for r_idx, row in enumerate(rows, start):
            bg = LGRAY if r_idx % 2 == 0 else WHITE
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = PatternFill(fill_type='solid', fgColor=bg)
                cell.font = Font(size=10)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.append(['LearnSync Manager Report'])
    ws['A1'].font = Font(bold=True, size=14, color=BLUE)
    ws.append(['Generated', data['today'].strftime('%d %B %Y')])
    ws.append(['Note', f"All metrics below exclude {data['unpaid_count']} learner(s) with Unpaid status."])
    ws['A3'].font = Font(bold=True, italic=True, size=10, color='B45309')
    ws.append([])
    ws.append(['Metric', 'Value', 'Definition'])
    for c in ws[5]:
        _hdr(c)
    for row in [
        ('Active Programmes',   data['active_programmes'],
         'Programmes currently running, excluding prerequisite/onboarding tracks.'),
        ('Paid Enrolled',       data['total_enrolled'],
         'Learners with a confirmed or pending payment enrolled in at least one programme.'),
        ('Activated',           data['total_activated'],
         'Enrolled learners who have passed the first module of their programme.'),
        ('Activation Rate',     f"{data['activation_rate']}%",
         'Activated ÷ Enrolled. What share of paid enrolled learners have passed the first module?'),
        ('Retained',            data['total_retained'],
         'Learners who passed the first module and remain engaged (health: Active, At Risk, or Graduated).'),
        ('Retention Rate',      f"{data['retention_rate']}%",
         'Retained ÷ Activated. Of those who passed the first module, how many are still going?'),
        ('Graduated',           data['total_graduated'],
         'Learners who completed all required courses and met graduation criteria.'),
        ('Graduation Rate',     f"{data['grad_rate']}%",
         'Graduated ÷ Enrolled. The end-to-end completion rate.'),
        ('Badges Earned',       data['total_badges'],
         'Total course-completion credentials awarded. One per passed course (learners can earn multiple).'),
        ('', '', ''),
        ('— Health Breakdown (paid only) —', '', ''),
        ('Active',      data['health_totals'].get('active', 0),
         'On track: submitting work, passing assignments, within inactivity thresholds.'),
        ('At Risk',     data['health_totals'].get('at_risk', 0),
         'At least one warning flag raised. See "At-Risk Flags" sheet for detail.'),
        ('Dormant',     data['health_totals'].get('dormant', 0),
         'Enrolled or activated but no meaningful activity for an extended period.'),
        ('Graduated',   data['health_totals'].get('graduated', 0),
         'Programme completed successfully.'),
        ('Not Started', data['health_totals'].get('not_yet_started', 0),
         'Paid and enrolled but no activity recorded yet.'),
    ]:
        ws.append(row)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 70

    # ── Sheet 2: By Programme ─────────────────────────────────────────────
    ws2 = wb.create_sheet('By Programme')
    prog_headers = [
        'Programme', 'Code',
        'Enrolled (paid)', 'Activated', 'Act %', 'Retained', 'Ret %',
        'Active', 'At Risk', 'Dormant', 'Not Started',
        'Graduated', 'Grad %', 'Badges',
    ]
    ws2.append(prog_headers)
    for cell in ws2[1]:
        _hdr(cell)
    _write_sheet_rows(ws2, [
        (
            p['programme__name'], p['programme__code'],
            p['total'], p['activated'], f"{p['activation_rate']}%",
            p['retained'], f"{p['retention_rate']}%",
            p['active'], p['at_risk'], p['dormant'], p['not_started'],
            p['graduated'], f"{p['grad_rate']}%", p['badges'],
        )
        for p in data['prog_rows']
    ])
    ws2.column_dimensions['A'].width = 40
    for i in range(2, len(prog_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: Learner Detail ───────────────────────────────────────────
    ws3 = wb.create_sheet('Learner Detail')
    learner_headers = [
        'First Name', 'Last Name', 'Email', 'Country', 'Payment Status',
        'Programme', 'Programme Name', 'Health Status', 'Current Course',
        'Enrolled Date', 'Activation Date', 'Graduated', 'Graduation Date',
    ]
    ws3.append(learner_headers)
    for cell in ws3[1]:
        _hdr(cell)
    pay_labels = dict(PaymentStatus.choices)
    _write_sheet_rows(ws3, [
        (
            r['learner__first_name'], r['learner__last_name'],
            r['learner__email'], r['learner__country'],
            pay_labels.get(r['learner__payment_status'], r['learner__payment_status']),
            r['programme__code'], r['programme__name'],
            _HEALTH_DISPLAY.get(r['health_status'], r['health_status']),
            r['current_course__full_name'] or '',
            str(r['enrolment_date'])   if r['enrolment_date']   else '',
            str(r['activation_date'])  if r['activation_date']  else '',
            'Yes' if r['is_graduated'] else 'No',
            str(r['graduation_date'])  if r['graduation_date']  else '',
        )
        for r in data['learner_rows']
    ])
    col_widths3 = [14, 14, 32, 14, 16, 12, 30, 14, 30, 14, 14, 12, 16]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A2'

    # ── Sheet 4: Payment Status ───────────────────────────────────────────
    ws4 = wb.create_sheet('Payment Status')
    ws4.append(['Payment Status', 'Learners', '% of All', 'Note'])
    for cell in ws4[1]:
        _hdr(cell)
    _write_sheet_rows(ws4, [
        (p['label'], p['count'], f"{p['pct']}%",
         'Excluded from performance metrics' if p['payment_status'] == PaymentStatus.UNKNOWN else '')
        for p in data['payment_rows']
    ])
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 34

    # ── Sheet 5: At-Risk Flags ────────────────────────────────────────────
    ws5 = wb.create_sheet('At-Risk Flags')
    ws5.append(['Flag', 'Paid Learners Affected', 'What it means'])
    for cell in ws5[1]:
        _hdr(cell)
    _write_sheet_rows(ws5, [(r['flag'], r['count'], r['desc']) for r in data['flag_rows']])
    ws5.column_dimensions['A'].width = 28
    ws5.column_dimensions['B'].width = 22
    ws5.column_dimensions['C'].width = 70

    # ── Sheet 6: Recent Interventions (30 days) ───────────────────────────
    ws6 = wb.create_sheet('Recent Interventions')
    ws6.append(['Date', 'Learner Email', 'Programme', 'Type', 'Outcome', 'Logged By', 'Notes'])
    for cell in ws6[1]:
        _hdr(cell)
    _write_sheet_rows(ws6, [
        (
            str(iv.intervention_date),
            iv.learner.email,
            iv.enrolment.programme.code if iv.enrolment else '',
            iv.get_type_display(),
            iv.get_outcome_display(),
            iv.logged_by.get_full_name() if iv.logged_by else '',
            iv.notes or '',
        )
        for iv in data['recent_interventions']
    ])
    col_widths6 = [14, 30, 14, 16, 20, 22, 40]
    for i, w in enumerate(col_widths6, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w
    ws6.freeze_panes = 'A2'

    # ── Sheet 7: Metric Definitions ───────────────────────────────────────
    ws7 = wb.create_sheet('Metric Definitions')
    ws7.append(['Metric', 'Definition'])
    for cell in ws7[1]:
        _hdr(cell)
    for i, (term, defn) in enumerate(data['metric_definitions'].items(), 2):
        bg = LGRAY if i % 2 == 0 else WHITE
        ws7.cell(row=i, column=1, value=term).font  = Font(bold=True, size=10)
        ws7.cell(row=i, column=1).fill = PatternFill(fill_type='solid', fgColor=bg)
        c = ws7.cell(row=i, column=2, value=defn)
        c.font = Font(size=10)
        c.fill = PatternFill(fill_type='solid', fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    ws7.column_dimensions['A'].width = 24
    ws7.column_dimensions['B'].width = 80
    ws7.row_dimensions[1].height = 18

    # ── Sheet 8: Weekly Breakdown ─────────────────────────────────────────
    ws8 = wb.create_sheet('Weekly Breakdown')
    ws8.append(['Week', 'Week Start', 'Week End', 'Enrolled', 'Activated', 'Active', 'At Risk', 'Graduated', 'Badges', 'Interventions'])
    for cell in ws8[1]:
        _hdr(cell)
    _write_sheet_rows(ws8, [
        (
            r['label'],
            str(r['week_start']),
            str(r['week_end']),
            r['enrolled'],
            r['activated'],
            r['active'],
            r['at_risk'],
            r['graduated'],
            r['badges'],
            r['interventions'],
        )
        for r in data['weekly_rows']
    ])
    col_widths8 = [20, 12, 12, 12, 12, 12, 12, 12, 12, 16]
    for i, w in enumerate(col_widths8, 1):
        ws8.column_dimensions[get_column_letter(i)].width = w
    ws8.freeze_panes = 'A2'

    # ── Sheet 9: Weekly by Programme ─────────────────────────────────────
    ws9 = wb.create_sheet('Weekly by Programme')
    ws9.append(['Programme', 'Code', 'Week', 'Enrolled', 'Activated', 'Active', 'At Risk', 'Graduated', 'Badges', 'Interventions'])
    for cell in ws9[1]:
        _hdr(cell)
    prog_week_rows = []
    for p in data['weekly_by_prog']:
        for w in p['weeks']:
            prog_week_rows.append((
                p['name'], p['code'], w['label'],
                w['enrolled'], w['activated'],
                w['active'], w['at_risk'],
                w['graduated'], w['badges'], w['interventions'],
            ))
    _write_sheet_rows(ws9, prog_week_rows)
    col_widths9 = [36, 10, 20, 12, 12, 12, 12, 12, 12, 16]
    for i, w in enumerate(col_widths9, 1):
        ws9.column_dimensions[get_column_letter(i)].width = w
    ws9.freeze_panes = 'A2'

    # ── Dashboard sheet (inserted at position 0) ─────────────────────────
    from openpyxl.chart import BarChart, LineChart, Reference

    ws_dash = wb.create_sheet('Dashboard', 0)
    ws_dash.sheet_properties.tabColor = BLUE

    # ── Title row ──────────────────────────────────────────────────────────
    ws_dash.merge_cells('B2:T2')
    ws_dash['B2'] = 'LearnSync — Manager Report Dashboard'
    ws_dash['B2'].font = Font(bold=True, size=16, color=BLUE)
    ws_dash['B2'].alignment = Alignment(vertical='center')
    ws_dash.row_dimensions[2].height = 28

    ws_dash.merge_cells('B3:T3')
    ws_dash['B3'] = (
        f"Generated {data['today'].strftime('%d %B %Y')}  ·  "
        f"Paid learners only  ·  {data['unpaid_count']} unpaid learner(s) excluded"
    )
    ws_dash['B3'].font = Font(size=10, color='6B7280')

    # ── KPI cards (row 5-7, 3 columns each) ───────────────────────────────
    kpi_cards = [
        (2,  'Active Programmes',   str(data['active_programmes']),                                    BLUE),
        (5,  'Paid Enrolled',       str(data['total_enrolled']),                                       '0D9488'),
        (8,  'Activated',           f"{data['total_activated']} ({data['activation_rate']}%)",         '7C3AED'),
        (11, 'Graduated',           f"{data['total_graduated']} ({data['grad_rate']}%)",               '16A34A'),
        (14, 'At Risk',             str(data['health_totals'].get('at_risk', 0)),                      'DC2626'),
        (17, 'Dormant / Not Started',
             f"{data['health_totals'].get('dormant', 0)} / {data['health_totals'].get('not_yet_started', 0)}",
             '6B7280'),
    ]
    ws_dash.row_dimensions[5].height = 16
    ws_dash.row_dimensions[6].height = 26
    ws_dash.row_dimensions[7].height = 6

    for start_col, label, value, color in kpi_cards:
        end_col = start_col + 2
        for row in (5, 6, 7):
            ws_dash.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=end_col,
            )
        lc = ws_dash.cell(row=5, column=start_col, value=label)
        lc.font = Font(bold=True, color=WHITE, size=9)
        lc.fill = PatternFill(fill_type='solid', fgColor=color)
        lc.alignment = Alignment(horizontal='center', vertical='center')

        vc = ws_dash.cell(row=6, column=start_col, value=value)
        vc.font = Font(bold=True, color=WHITE, size=16)
        vc.fill = PatternFill(fill_type='solid', fgColor=color)
        vc.alignment = Alignment(horizontal='center', vertical='center')

        pc = ws_dash.cell(row=7, column=start_col)
        pc.fill = PatternFill(fill_type='solid', fgColor=color)

        for c in range(start_col, end_col + 1):
            ws_dash.column_dimensions[get_column_letter(c)].width = 7

    # ── Chart data (written below row 50, used as chart data sources) ─────
    # Health breakdown data  (row 52–57)
    HD_ROW = 52
    ws_dash.cell(row=HD_ROW,     column=1, value='Health Status')
    ws_dash.cell(row=HD_ROW,     column=2, value='Learners')
    health_items = [
        ('Active',      data['health_totals'].get('active', 0)),
        ('At Risk',     data['health_totals'].get('at_risk', 0)),
        ('Dormant',     data['health_totals'].get('dormant', 0)),
        ('Graduated',   data['health_totals'].get('graduated', 0)),
        ('Not Started', data['health_totals'].get('not_yet_started', 0)),
    ]
    for i, (lbl, val) in enumerate(health_items, 1):
        ws_dash.cell(row=HD_ROW + i, column=1, value=lbl)
        ws_dash.cell(row=HD_ROW + i, column=2, value=val)

    # Programme comparison data (row 62+)
    PD_ROW = 62
    ws_dash.cell(row=PD_ROW, column=1, value='Programme')
    ws_dash.cell(row=PD_ROW, column=2, value='Enrolled')
    ws_dash.cell(row=PD_ROW, column=3, value='Activated')
    ws_dash.cell(row=PD_ROW, column=4, value='Graduated')
    for i, p in enumerate(data['prog_rows'], 1):
        ws_dash.cell(row=PD_ROW + i, column=1, value=p['programme__code'])
        ws_dash.cell(row=PD_ROW + i, column=2, value=p['total'])
        ws_dash.cell(row=PD_ROW + i, column=3, value=p['activated'])
        ws_dash.cell(row=PD_ROW + i, column=4, value=p['graduated'])

    # Weekly trend data (row 80+)
    WD_ROW = 80
    ws_dash.cell(row=WD_ROW, column=1, value='Week')
    ws_dash.cell(row=WD_ROW, column=2, value='Enrolled')
    ws_dash.cell(row=WD_ROW, column=3, value='Activated')
    ws_dash.cell(row=WD_ROW, column=4, value='Active')
    ws_dash.cell(row=WD_ROW, column=5, value='At Risk')
    ws_dash.cell(row=WD_ROW, column=6, value='Graduated')
    for i, r in enumerate(data['weekly_rows'], 1):
        ws_dash.cell(row=WD_ROW + i, column=1, value=r['label'])
        ws_dash.cell(row=WD_ROW + i, column=2, value=r['enrolled'])
        ws_dash.cell(row=WD_ROW + i, column=3, value=r['activated'])
        ws_dash.cell(row=WD_ROW + i, column=4, value=r['active'])
        ws_dash.cell(row=WD_ROW + i, column=5, value=r['at_risk'])
        ws_dash.cell(row=WD_ROW + i, column=6, value=r['graduated'])

    # At-risk flag data (row 97+)
    FD_ROW = 97
    ws_dash.cell(row=FD_ROW, column=1, value='Flag')
    ws_dash.cell(row=FD_ROW, column=2, value='Learners')
    for i, f in enumerate(data['flag_rows'], 1):
        ws_dash.cell(row=FD_ROW + i, column=1, value=f['flag'])
        ws_dash.cell(row=FD_ROW + i, column=2, value=f['count'])

    # ── Chart 1: Health status horizontal bar ─────────────────────────────
    hc = BarChart()
    hc.type   = 'bar'   # horizontal
    hc.title  = 'Health Status Breakdown'
    hc.y_axis.title = 'Status'
    hc.x_axis.title = 'Learners'
    hc.height = 10
    hc.width  = 17
    hc.legend.position = 'b'
    hc_data = Reference(ws_dash, min_col=2, min_row=HD_ROW,
                         max_row=HD_ROW + len(health_items))
    hc_cats = Reference(ws_dash, min_col=1, min_row=HD_ROW + 1,
                         max_row=HD_ROW + len(health_items))
    hc.add_data(hc_data, titles_from_data=True)
    hc.set_categories(hc_cats)
    ws_dash.add_chart(hc, 'B9')

    # ── Chart 2: Programme comparison grouped bar ─────────────────────────
    if data['prog_rows']:
        pc2 = BarChart()
        pc2.type      = 'col'   # vertical grouped
        pc2.grouping  = 'clustered'
        pc2.title     = 'Programme: Enrolled / Activated / Graduated'
        pc2.y_axis.title = 'Learners'
        pc2.height    = 10
        pc2.width     = 22
        n_prog = len(data['prog_rows'])
        pc2_data = Reference(ws_dash, min_col=2, max_col=4,
                              min_row=PD_ROW, max_row=PD_ROW + n_prog)
        pc2_cats = Reference(ws_dash, min_col=1,
                              min_row=PD_ROW + 1, max_row=PD_ROW + n_prog)
        pc2.add_data(pc2_data, titles_from_data=True)
        pc2.set_categories(pc2_cats)
        ws_dash.add_chart(pc2, 'L9')

    # ── Chart 3: Weekly trend line ─────────────────────────────────────────
    if data['weekly_rows']:
        wc = LineChart()
        wc.title   = 'Weekly Activity Trend (last 13 weeks)'
        wc.y_axis.title = 'Learners'
        wc.height  = 12
        wc.width   = 34
        n_weeks = len(data['weekly_rows'])
        wc_data = Reference(ws_dash, min_col=2, max_col=6,
                             min_row=WD_ROW, max_row=WD_ROW + n_weeks)
        wc_cats = Reference(ws_dash, min_col=1,
                             min_row=WD_ROW + 1, max_row=WD_ROW + n_weeks)
        wc.add_data(wc_data, titles_from_data=True)
        wc.set_categories(wc_cats)
        wc.smooth = True
        ws_dash.add_chart(wc, 'B24')

    # ── Chart 4: At-risk flags horizontal bar ─────────────────────────────
    if data['flag_rows']:
        fc = BarChart()
        fc.type   = 'bar'
        fc.title  = 'At-Risk Flags — Learners Affected'
        fc.x_axis.title = 'Learners'
        fc.height = 10
        fc.width  = 17
        n_flags = len(data['flag_rows'])
        fc_data = Reference(ws_dash, min_col=2, min_row=FD_ROW,
                             max_row=FD_ROW + n_flags)
        fc_cats = Reference(ws_dash, min_col=1,
                             min_row=FD_ROW + 1, max_row=FD_ROW + n_flags)
        fc.add_data(fc_data, titles_from_data=True)
        fc.set_categories(fc_cats)
        ws_dash.add_chart(fc, 'L24')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"learnsync_report_{data['today'].isoformat()}.xlsx"
    resp = HttpResponse(buf.read(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _pdf_response(data):
    from reportlab.graphics.charts.barcharts import HorizontalBarChart, VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    BLUE  = colors.HexColor('#0452F0')
    LIGHT = colors.HexColor('#EFF6FF')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Heading1'],
                             textColor=BLUE, fontSize=18, spaceAfter=4)
    sub_s   = ParagraphStyle('S', parent=styles['Normal'],
                             textColor=colors.grey, fontSize=10, spaceAfter=16)
    sec_s   = ParagraphStyle('H', parent=styles['Heading2'],
                             textColor=BLUE, fontSize=12, spaceBefore=14, spaceAfter=6)

    def _ts():
        return TableStyle([
            ('BACKGROUND',    (0, 0), (-1,  0), BLUE),
            ('TEXTCOLOR',     (0, 0), (-1,  0), colors.white),
            ('FONTNAME',      (0, 0), (-1,  0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, LIGHT]),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
            ('ALIGN',         (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN',         (0, 0), (0,  -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (0,  -1), 8),
        ])

    # Usable page width on landscape A4 minus margins
    W = landscape(A4)[0] - 3*cm

    # ── Chart helpers ──────────────────────────────────────────────────────
    def _hbar_chart(labels, values, bar_colors, height=130):
        """Single-series horizontal bar chart — one colour per bar."""
        d = Drawing(W, height)
        bc = HorizontalBarChart()
        bc.x, bc.y   = 75, 8
        bc.width     = W - 90
        bc.height    = height - 16
        bc.data      = [list(values)]
        bc.categoryAxis.categoryNames = list(labels)
        bc.categoryAxis.labels.fontSize  = 8
        bc.categoryAxis.labels.dx        = -3
        bc.categoryAxis.tickShift        = True
        bc.valueAxis.labels.fontSize     = 8
        bc.valueAxis.forceZero           = 1
        bc.bars.strokeWidth              = 0
        for i, c in enumerate(bar_colors):
            bc.bars[0, i].fillColor = colors.HexColor(c)
        d.add(bc)
        return d

    def _vbar_grouped_chart(cat_labels, series_data, series_labels, series_colors, height=200):
        """Multi-series vertical grouped bar chart."""
        d = Drawing(W, height)
        bc = VerticalBarChart()
        bc.x, bc.y = 35, 30
        bc.width   = W - 55
        bc.height  = height - 55
        bc.data    = [list(s) for s in series_data]
        bc.categoryAxis.categoryNames = [str(l)[:14] for l in cat_labels]
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.angle    = 30 if len(cat_labels) > 5 else 0
        bc.categoryAxis.labels.dy       = -8 if len(cat_labels) > 5 else 0
        bc.valueAxis.labels.fontSize    = 7
        bc.valueAxis.forceZero          = 1
        bc.groupSpacing                 = 5
        bc.bars.strokeWidth             = 0
        for i, c in enumerate(series_colors):
            bc.bars[i].fillColor = colors.HexColor(c)
        # Legend
        lg = Legend()
        lg.x, lg.y    = bc.x, height - 18
        lg.fontSize   = 8
        lg.columnMaximum = 1
        lg.dy         = -12
        lg.dx         = 60
        lg.colorNamePairs = [(colors.HexColor(c), lbl)
                             for c, lbl in zip(series_colors, series_labels)]
        d.add(bc)
        d.add(lg)
        return d

    def _line_chart(x_labels, series_data, series_labels, series_colors, height=170):
        """Multi-series horizontal line chart."""
        d = Drawing(W, height)
        lc = HorizontalLineChart()
        lc.x, lc.y  = 35, 28
        lc.width    = W - 55
        lc.height   = height - 50
        lc.data     = [list(s) for s in series_data]
        lc.joinedLines    = 1
        lc.categoryAxis.categoryNames = [str(l)[:8] for l in x_labels]
        lc.categoryAxis.labels.fontSize = 6
        lc.categoryAxis.labels.angle    = 40
        lc.categoryAxis.labels.dy       = -8
        lc.valueAxis.labels.fontSize    = 7
        lc.valueAxis.forceZero          = 1
        for i, c in enumerate(series_colors):
            lc.lines[i].strokeColor = colors.HexColor(c)
            lc.lines[i].strokeWidth = 1.8
        lg = Legend()
        lg.x, lg.y    = lc.x, height - 18
        lg.fontSize   = 8
        lg.dx         = 60
        lg.dy         = -12
        lg.columnMaximum = 1
        lg.colorNamePairs = [(colors.HexColor(c), lbl)
                             for c, lbl in zip(series_colors, series_labels)]
        d.add(lc)
        d.add(lg)
        return d

    note_s  = ParagraphStyle('N', parent=styles['Normal'],
                             textColor=colors.HexColor('#92400E'),
                             backColor=colors.HexColor('#FEF3C7'),
                             fontSize=8, spaceAfter=12, leading=12,
                             leftIndent=6, rightIndent=6)

    story = [
        Paragraph('LearnSync Manager Report', title_s),
        Paragraph(f"Generated {data['today'].strftime('%d %B %Y')} &nbsp;·&nbsp; Paid learners only (excl. {data['unpaid_count']} unpaid)", sub_s),
        Paragraph('Performance Summary', sec_s),
    ]

    summary_data = [
        ['Metric', 'Value', 'What it means'],
        ['Active Programmes',   str(data['active_programmes']),
         'Programmes currently running, excluding prerequisite tracks.'],
        ['Paid Enrolled',       str(data['total_enrolled']),
         'Enrolled learners with confirmed or pending payment.'],
        ['Activated',           f"{data['total_activated']} ({data['activation_rate']}%)",
         'Learners who have passed the first module of their programme. Rate = Activated ÷ Enrolled.'],
        ['Retained',            f"{data['total_retained']} ({data['retention_rate']}%)",
         'Passed first module and remain engaged. Rate = Retained ÷ Activated.'],
        ['Graduated',           f"{data['total_graduated']} ({data['grad_rate']}%)",
         'Completed all required courses. Rate = Graduated ÷ Enrolled.'],
        ['Badges Earned',       str(data['total_badges']),
         'Course-completion credentials awarded. One per passed course.'],
    ]
    t1 = Table(summary_data, colWidths=[W*0.22, W*0.15, W*0.63])
    t1.setStyle(_ts())
    story += [t1, Spacer(1, 0.4*cm)]

    # Health breakdown
    story.append(Paragraph('Health Status Overview  (paid learners only)', sec_s))
    health_rows = [
        ('active',          'Active',      'On track — submitting, passing, within thresholds.'),
        ('at_risk',         'At Risk',     'One or more warning flags raised.'),
        ('dormant',         'Dormant',     'No meaningful activity for an extended period.'),
        ('graduated',       'Graduated',   'Programme completed.'),
        ('not_yet_started', 'Not Started', 'Paid and enrolled but no activity yet.'),
    ]
    health_data = [['Status', 'Learners', 'What it means']] + [
        [label, str(data['health_totals'].get(k, 0)), desc]
        for k, label, desc in health_rows
    ]
    t_h = Table(health_data, colWidths=[W*0.18, W*0.12, W*0.70])
    t_h.setStyle(_ts())
    story.append(t_h)
    # Health bar chart — bottom-to-top order so Active appears at the top
    story.append(_hbar_chart(
        labels=['Not Started', 'Graduated', 'Dormant', 'At Risk', 'Active'],
        values=[
            data['health_totals'].get('not_yet_started', 0),
            data['health_totals'].get('graduated', 0),
            data['health_totals'].get('dormant', 0),
            data['health_totals'].get('at_risk', 0),
            data['health_totals'].get('active', 0),
        ],
        bar_colors=['#9ca3af', '#16a34a', '#6b7280', '#dc2626', '#0d9488'],
    ))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph('Programme Breakdown  (paid learners only)', sec_s))
    pw = [W*0.20, W*0.07, W*0.07, W*0.07, W*0.07, W*0.07, W*0.07, W*0.07, W*0.08, W*0.08, W*0.07, W*0.08]
    prog_data = [['Programme', 'Enrolled', 'Act\'d', 'Act %', 'Ret\'d', 'Ret %',
                  'Active', 'At Risk', 'Dormant', 'Grad\'d', 'Grad %', 'Badges']] + [
        [p['programme__name'][:30],
         str(p['total']),
         str(p['activated']),  f"{p['activation_rate']}%",
         str(p['retained']),   f"{p['retention_rate']}%",
         str(p['active']),     str(p['at_risk']),  str(p['dormant']),
         str(p['graduated']),  f"{p['grad_rate']}%", str(p['badges'])]
        for p in data['prog_rows']
    ]
    t2 = Table(prog_data, colWidths=pw, repeatRows=1)
    t2.setStyle(_ts())
    story.append(t2)
    # Programme grouped bar chart — Enrolled / Activated / Graduated per programme
    if data['prog_rows']:
        story.append(_vbar_grouped_chart(
            cat_labels=[p['programme__code'] for p in data['prog_rows']],
            series_data=[
                [p['total']     for p in data['prog_rows']],
                [p['activated'] for p in data['prog_rows']],
                [p['graduated'] for p in data['prog_rows']],
            ],
            series_labels=['Enrolled', 'Activated', 'Graduated'],
            series_colors=['#0452F0', '#7c3aed', '#16a34a'],
        ))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph('At-Risk Flag Breakdown', sec_s))
    flag_data = [['Flag', 'Paid Learners', 'What it means']] + [
        [r['flag'], str(r['count']), r['desc']]
        for r in data['flag_rows']
    ]
    t_f = Table(flag_data, colWidths=[W*0.22, W*0.10, W*0.68], repeatRows=1)
    t_f.setStyle(_ts())
    story += [t_f, Spacer(1, 0.4*cm)]

    story.append(Paragraph('Payment Status Distribution  (all learners including unpaid)', sec_s))
    pay_data = [['Status', 'Learners', '%', 'Note']] + [
        [p['label'], str(p['count']), f"{p['pct']}%",
         'Excluded from performance metrics' if p['payment_status'] == PaymentStatus.UNKNOWN else '']
        for p in data['payment_rows']
    ]
    t3 = Table(pay_data, colWidths=[W*0.22, W*0.12, W*0.10, W*0.56])
    t3.setStyle(_ts())
    story.append(t3)

    story += [Spacer(1, 0.4*cm)]
    story.append(Paragraph('Weekly Activity  (last 13 weeks, paid learners)', sec_s))
    weekly_data = [['Week', 'Enrolled', 'Activated', 'Active', 'At Risk', 'Graduated', 'Interventions']] + [
        [r['label'], str(r['enrolled']), str(r['activated']),
         str(r['active']), str(r['at_risk']), str(r['graduated']), str(r['interventions'])]
        for r in data['weekly_rows']
    ]
    t_w = Table(weekly_data, colWidths=[W*0.24, W*0.12, W*0.12, W*0.12, W*0.12, W*0.12, W*0.16], repeatRows=1)
    t_w.setStyle(_ts())
    story.append(t_w)
    # Weekly line chart — Enrolled / Activated / Graduated trend
    if data['weekly_rows']:
        _wk_labels = [r['label'].split('–')[0].strip() for r in data['weekly_rows']]
        story.append(_line_chart(
            x_labels=_wk_labels,
            series_data=[
                [r['enrolled']  for r in data['weekly_rows']],
                [r['activated'] for r in data['weekly_rows']],
                [r['active']    for r in data['weekly_rows']],
                [r['at_risk']   for r in data['weekly_rows']],
                [r['graduated'] for r in data['weekly_rows']],
            ],
            series_labels=['Enrolled', 'Activated', 'Active', 'At Risk', 'Graduated'],
            series_colors=['#0452F0', '#7c3aed', '#0d9488', '#dc2626', '#16a34a'],
        ))

    doc.build(story)
    buf.seek(0)
    filename = f"learnsync_report_{data['today'].isoformat()}.pdf"
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
